import os
import time
import json
from datetime import datetime
from pathlib import Path
import pandas as pd
from pdfminer.high_level import extract_text
from pdfminer.layout import LAParams
import re
from openpyxl import load_workbook, Workbook
 
 
# ============================================================
# PDF MINER FUNCTION - SWITCH TO OCR IF NECESSARY
# ============================================================
def extract_pdf_text(pdf_path):
    """
    Extract text from PDF using pdfminer (NOT OCR YET).
    If switching to Microsoft Azure OCR, replace this function body.
 
    Args:
        pdf_path: Path to PDF file
 
    Returns:
        Extracted text as string
    """
    # RUBEN: Be careful of Microsoft Azure implementation in Python — can cause bugs
    laparams = LAParams(
        line_margin=0.5,
        char_margin=2.0,
        word_margin=10.0,   # Large to capture right-aligned values if possible
        boxes_flow=0.5,
        detect_vertical=False,
        all_texts=True
    )
    text = extract_text(pdf_path, laparams=laparams)
    return text
# END PDF MINER FUNCTION
 
 
# ============================================================
# KEYWORD SEARCH
# ============================================================
def extract_values_after_keywords(pdf_path, keyword_offset_dict):
    """
    Extract values X lines after keywords from the PDF-mined text.
    Each keyword can have its own specific line offset to account for PDF irregularity.
    Handles keywords split across two consecutive lines (Method 2).
 
    FIX: Method 2 (split-line search) is now correctly nested inside the
    per-keyword loop so it runs for every keyword, not just the last one.
 
    Args:
        pdf_path:            Path to PDF file
        keyword_offset_dict: Dict mapping keywords to their line offsets
                             e.g. {'Date': 2, 'Amount Due': 2, 'Meter Number': 22}
 
    Returns:
        DataFrame with all raw extraction results (may contain multiple rows
        per keyword — call select_best_result() to reduce to one row each).
    """
    text  = extract_pdf_text(pdf_path)
    lines = text.split('\n')
    results         = []
    integer_pattern = r'\b\d+\b'
 
    for keyword, lines_offset in keyword_offset_dict.items():
        keyword_no_spaces = keyword.replace(' ', '')
        pattern           = re.compile(re.escape(keyword), re.IGNORECASE)
 
        # ------------------------------------------------------------------
        # Method 1: keyword found on a single line
        # ------------------------------------------------------------------
        for line_num, line in enumerate(lines):
            match = pattern.search(line)
            if match:
                target_line_num = line_num + lines_offset
                if target_line_num < len(lines):
                    target_line = lines[target_line_num].strip()
 
                    if keyword.lower() in ['ccf', 'kwh']:
                        int_match       = re.search(integer_pattern, target_line)
                        extracted_value = int_match.group() if int_match else ""
                    else:
                        extracted_value = target_line
                else:
                    extracted_value = ""
 
                if extracted_value:
                    results.append({
                        'keyword':         keyword,
                        'found_at_line':   line_num,
                        'line_offset':     lines_offset,
                        'target_line':     target_line_num,
                        'extracted_value': extracted_value,
                        'match_type':      'single_line'
                    })
 
        # ------------------------------------------------------------------
        # Method 2: keyword split across two consecutive lines
        # (FIX) This block was previously OUTSIDE the keyword loop, so it
        # only ever ran for the last keyword in the dict.  It is now
        # correctly indented inside the loop so every keyword is checked.
        # ------------------------------------------------------------------
        for line_num in range(len(lines) - 1):
            combined = (lines[line_num] + lines[line_num + 1]).replace(' ', '').replace('\n', '')
 
            if keyword_no_spaces.lower() in combined.lower():
                # Keyword ends on line_num+1, so offset starts from there
                target_line_num = (line_num + 1) + lines_offset
 
                if target_line_num < len(lines):
                    target_line = lines[target_line_num].strip()
 
                    if keyword.lower() in ['ccf', 'kwh']:
                        int_match       = re.search(integer_pattern, target_line)
                        extracted_value = int_match.group() if int_match else ""
                    else:
                        extracted_value = target_line
                else:
                    extracted_value = ""
 
                if extracted_value:
                    results.append({
                        'keyword':         keyword,
                        'found_at_line':   f"{line_num}-{line_num+1} (split)",
                        'line_offset':     lines_offset,
                        'target_line':     target_line_num,
                        'extracted_value': extracted_value,
                        'match_type':      'split_lines'
                    })
 
        # ------------------------------------------------------------------
        # Method 3: regex fallback for Meter Number only
        # Triggered when Methods 1 & 2 found nothing with exactly 10 digits.
        # Scans every line in the document for a line containing ONLY a
        # 10-digit integer (no letters or symbols).
        # ------------------------------------------------------------------
        if keyword.lower() == 'meter number':
            # Check if any result so far is actually a valid 10-digit number
            existing = [
                r['extracted_value'] for r in results
                if r['keyword'] == keyword
            ]
            has_valid = any(
                len(re.sub(r'\D', '', v)) == 10 for v in existing
            )

            if not has_valid:
                print(f"  → Meter Number: offset search failed, trying regex fallback...")
                ten_digit_pattern = re.compile(r'^\d{10}$')
                for line_num, line in enumerate(lines):
                    stripped = line.strip()
                    if ten_digit_pattern.match(stripped):
                        results.append({
                            'keyword':         keyword,
                            'found_at_line':   line_num,
                            'line_offset':     'regex_fallback',
                            'target_line':     line_num,
                            'extracted_value': stripped,
                            'match_type':      'regex_fallback'
                        })
                        print(f"  → Regex fallback found: {stripped} at line {line_num}")
                        break  # Take the first match and stop

    df = pd.DataFrame(results)
    if not df.empty:
        df = df.drop_duplicates(subset=['keyword', 'extracted_value'], keep='first')

    return df
# END KEYWORD SEARCH
 
 
# ============================================================
# BEST-RESULT SELECTION
# ============================================================
def _parse_dollar(value_str):
    """
    Parse a dollar-amount string into a float.
    Strips $, commas, and surrounding whitespace.
    Returns None if the string cannot be parsed as a number.
    """
    cleaned = re.sub(r'[^\d.]', '', str(value_str).replace(',', ''))
    try:
        return float(cleaned)
    except ValueError:
        return None
 
 
def select_best_result(keyword, candidates):
    """
    Given a keyword and a list of candidate extracted_value strings,
    return the single best value according to per-keyword business rules:
 
        'amount due'              → largest dollar amount
        'statement date'          → first (earliest-encountered) value
        'meter number'            → first value that is exactly 10 digits
        'kwh'                     → largest integer
        'ccf'                     → largest integer
        'total electricity cost'  → largest dollar amount
        'total natural gas cost'  → largest dollar amount
 
    Falls back to the first candidate if no rule produces a match.
 
    Args:
        keyword:    The keyword string (case-insensitive match used internally)
        candidates: List of extracted_value strings for this keyword
 
    Returns:
        Best value string, or "" if candidates is empty.
    """
    if not candidates:
        return ""
 
    kw = keyword.lower().strip()
 
    # --- largest dollar amount -------------------------------------------
    if kw in ('amount due', 'total electricity cost', 'total natural gas cost'):
        parsed = [(v, _parse_dollar(v)) for v in candidates]
        valid  = [(v, amt) for v, amt in parsed if amt is not None]
        if valid:
            return max(valid, key=lambda x: x[1])[0]
 
    # --- first date value ------------------------------------------------
    elif kw == 'statement date':
        return candidates[0]
 
    # --- exactly 10-digit integer ----------------------------------------
    elif kw == 'meter number':
        for v in candidates:
            digits_only = re.sub(r'\D', '', v)
            if len(digits_only) == 10:
                return digits_only          # return clean digits
        # Fallback: return first candidate if none are exactly 10 digits
        return candidates[0]
 
    # --- largest integer -------------------------------------------------
    elif kw in ('kwh', 'ccf'):
        parsed = []
        for v in candidates:
            m = re.search(r'\d+', v)
            if m:
                parsed.append((v, int(m.group())))
        if parsed:
            return max(parsed, key=lambda x: x[1])[0]
 
    # --- default: first candidate ----------------------------------------
    return candidates[0]
 
 
def build_single_row(df_raw, keyword_offset_dict):
    """
    Reduce the raw multi-row extraction DataFrame to a single output row
    (one best value per keyword) in the same order as keyword_offset_dict.
 
    Args:
        df_raw:              DataFrame returned by extract_values_after_keywords()
        keyword_offset_dict: Ordered dict of keywords (used for column ordering)
 
    Returns:
        Single-row DataFrame with one column per keyword (keyword as column name).
    """
    row = {}
    for keyword in keyword_offset_dict:
        if not df_raw.empty and keyword in df_raw['keyword'].values:
            candidates = (
                df_raw[df_raw['keyword'] == keyword]['extracted_value']
                .dropna()
                .tolist()
            )
            row[keyword] = select_best_result(keyword, candidates)
        else:
            row[keyword] = ""
 
    return pd.DataFrame([row])
# END BEST-RESULT SELECTION
 
 
# ============================================================
# AUTOMATED PDF SCANNING
# Redesigned for single-run / button-triggered execution.
# Call scanner.run_scan(keyword_offset_dict) once per button press.
# ============================================================
class PDFScanner:
    def __init__(self, watch_folder, output_folder, processed_log='processed_files.json'):
        """
        Args:
            watch_folder:   Folder to monitor for new PDFs
            output_folder:  Folder to save extraction results / debug files
            processed_log:  JSON file to track already-processed files
        """
        self.watch_folder  = Path(watch_folder)
        self.output_folder = Path(output_folder)
        self.processed_log = processed_log
 
        self.output_folder.mkdir(exist_ok=True)
        self.processed_files = self._load_processed_log()
 
    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------
    def _load_processed_log(self):
        if os.path.exists(self.processed_log):
            with open(self.processed_log, 'r') as f:
                return json.load(f)
        return {}
 
    def _save_processed_log(self):
        with open(self.processed_log, 'w') as f:
            json.dump(self.processed_files, f, indent=2)
 
    def _get_new_files(self):
        """Return PDFs in watch_folder that are new or modified since last run."""
        all_pdfs  = list(self.watch_folder.glob('*.pdf'))
        new_files = []
        for pdf_file in all_pdfs:
            file_key      = str(pdf_file)
            file_modified = os.path.getmtime(pdf_file)
            if file_key not in self.processed_files or \
               self.processed_files[file_key] != file_modified:
                new_files.append(pdf_file)
        return new_files
 
    def _process_single_file(self, pdf_path, keyword_offset_dict):
        """
        Extract raw keyword data from one PDF and return the best-result row.
 
        Args:
            pdf_path:            Path object for the PDF
            keyword_offset_dict: Keyword → offset mapping
 
        Returns:
            Single-row DataFrame, or None on error.
        """
        print(f"\n{'='*60}")
        print(f"Processing: {pdf_path.name}")
        print(f"{'='*60}")
 
        try:
            # Save raw text for debugging
            # TODO: remove _raw.txt output in final implementation
            full_text       = extract_pdf_text(str(pdf_path))
            txt_output_file = self.output_folder / f"{pdf_path.stem}_raw.txt"
            with open(txt_output_file, 'w', encoding='utf-8') as f:
                f.write(full_text)
            print(f"✓ Saved raw text to {txt_output_file}")
 
            # Extract all candidate values
            df_raw = extract_values_after_keywords(str(pdf_path), keyword_offset_dict)
            df_raw['source_file']    = pdf_path.name
            df_raw['processed_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
 
            print(f"✓ Raw extraction: {len(df_raw)} candidate values found")
 
            # Reduce to one best value per keyword
            df_row = build_single_row(df_raw, keyword_offset_dict)
            df_row['source_file']    = pdf_path.name
            df_row['processed_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
 
            print(f"✓ Best-result row built ({len(df_row.columns)} columns)")
            return df_row
 
        except Exception as e:
            print(f"✗ Error processing {pdf_path.name}: {e}")
            return None
 
    # ------------------------------------------------------------------
    # Public entry point — call this once per button press
    # ------------------------------------------------------------------
    def run_scan(self, keyword_offset_dict, excel_file_path):
        """
        Scan the watch folder for new PDFs, extract data, and append
        each result as a new row in the target Excel file.
 
        Designed to be called once per button press (no loop / sleep).
 
        Args:
            keyword_offset_dict: Keyword → line-offset mapping
            excel_file_path:     Absolute path to the existing target Excel file
        """
        print(f"\n{'#'*60}")
        print(f"Scan triggered at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Watching: {self.watch_folder}")
        print(f"{'#'*60}")
 
        new_files = self._get_new_files()
 
        if not new_files:
            print("✓ No new files found — nothing to process.")
            return
 
        print(f"✓ {len(new_files)} new file(s) to process")
 
        excel_path = Path(excel_file_path)
        if not excel_path.exists():
            print(f"✗ Target Excel file not found: {excel_path}")
            return
 
        for pdf_file in new_files:
            df_row = self._process_single_file(pdf_file, keyword_offset_dict)
 
            if df_row is not None and not df_row.empty:
                try:
                    book       = load_workbook(excel_path)
                    sheet_name = "Input Data"
 
                    if sheet_name not in book.sheetnames:
                        print(f"✗ Sheet '{sheet_name}' not found in {excel_path.name}")
                        continue
 
                    sheet     = book[sheet_name]
                    start_row = sheet.max_row  # Append after last used row
 
                    with pd.ExcelWriter(
                        excel_path,
                        engine="openpyxl",
                        mode="a",
                        if_sheet_exists="overlay"
                    ) as writer:
                        df_row.to_excel(
                            writer,
                            sheet_name=sheet_name,
                            index=False,
                            header=False,       # Don't re-write headers
                            startrow=start_row
                        )
 
                    print(f"✓ Appended row to '{sheet_name}' in {excel_path.name}")
 
                    # Mark as processed only after a successful write
                    self.processed_files[str(pdf_file)] = os.path.getmtime(pdf_file)
 
                except Exception as e:
                    print(f"✗ Failed to write {pdf_file.name} to Excel: {e}")
 
        self._save_processed_log()
        print(f"\n✓ Scan complete!")
# END AUTOMATED PDF SCANNING
 
 
# ============================================================
# USAGE — called when a button triggers this script
# ============================================================
if __name__ == "__main__":
    WATCH_FOLDER  = 'incoming_pdfs'     # Folder to monitor for new PDFs
    OUTPUT_FOLDER = 'extracted_data'    # Where to save debug/intermediate files
    EXCEL_FILE    = '/Users/ecmerritt/Desktop/Capstone/Testing.xlsx'
 
    # Keyword → line offset dictionary
    # Using RGE 32 as reference
    KEYWORD_OFFSETS = {
        'Amount Due':             2,
        'Statement Date':         2,
        'Meter Number':           22,
        'kwh':                    0,   # Special export: integer only
        'ccf':                    0,   # Special export: integer only
        'Total Electricity Cost': 38,
        'Total Natural Gas Cost': 15,
    }
 
    # ------------------------------------------------------------------
    # This block is the single-run entry point.
    # Replace with a button callback when I work out the UI portion:) —
    # ------------------------------------------------------------------
    scanner = PDFScanner(WATCH_FOLDER, OUTPUT_FOLDER)
    scanner.run_scan(KEYWORD_OFFSETS, EXCEL_FILE)
  
